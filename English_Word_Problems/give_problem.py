import openpyxl
import glob
import os
import random

import scraping


def main():
    # 出題用のデータファイルが存在するかを調べる
    file_name: str = '\\English_Word.xlsx'
    file_dir: str = __file__.removesuffix('\\' + os.path.basename(__file__))
    file_path: str = file_dir + file_name

    files: list = glob.glob(file_dir + '\\*')

    flag: bool = False
    for f in files:
        if f == (file_path):
            print('ファイルが存在します')
            flag = True
            break

    if flag == False:
        scraping.create_data()

    # 出題用英単語を取得（listに格納）
    num = int(input('出題数を入力してください：'))
    print()

    words: list = []
    i: int = 1
    while len(words) != num:
        # 出題用の単語を取得
        word: tuple = get_engword(file_path)
        fake_word: list = get_fakeword(word[0], file_path)
        random.shuffle(fake_word)

        if word[1] not in words:
            # 出題
            print(f'第{i}問')
            print(word[1])

            print(f'1.{fake_word[0]} 2.{fake_word[1]} 3.{fake_word[2]} 4.{fake_word[3]}')

            input_ans: str = input('答えの番号を入力してください -> ')

            # 解答のチェック
            if fake_word[int(input_ans)-1] == get_ans(word[0], file_path):
                print('正解')
            else:
                print('不正解')

            print()

            words.append(word)
            i += 1


# ランダムでファイルから英単語を取得する
def get_engword(file_path: str) -> tuple:
    wb = openpyxl.load_workbook(file_path)
    wb.active
    ws = wb['word_data']

    max: int = ws.max_row

    # 乱数生成
    rnd: int = random.randint(1, max)
    wb.save(file_path)

    return rnd, ws['A' + str(rnd)].value


# 選択肢を取得する
def get_fakeword(ans_word: int, file_path: str) -> list:
    wb = openpyxl.load_workbook(file_path)
    wb.active
    ws = wb['word_data']

    max: int = ws.max_row

    fake_list = []
    while len(fake_list) != 3:
        rnd = random.randint(1, max)

        if rnd != ans_word:
            fake_list.append(rnd)
        else:
            continue

    fake_word = []    
    for f in fake_list:
        fake_word.append(ws['B' + str(f)].value)

    fake_word.append(ws['B' + str(ans_word)].value)

    wb.save(file_path)
    return fake_word


# 解答を取得する
def get_ans(ans_word: int, file_path: str) -> str:
    wb = openpyxl.load_workbook(file_path)
    wb.active
    ws = wb['word_data']

    ans = ws['B' + str(ans_word)].value
    wb.save(file_path)

    return ans


if __name__ == '__main__':
    main()