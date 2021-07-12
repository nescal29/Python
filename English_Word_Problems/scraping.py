import openpyxl
import requests
import bs4
from bs4 import BeautifulSoup

class create_data_file():
    def create_data():
        # Webページの取得
        page_url: str = 'https://toiguru.jp/toeic-vocabulary-list'
        page: requests.models.Response = requests.get(page_url)

        soup: bs4.BeautifulSoup = BeautifulSoup(page.text, 'lxml')

        # tbタグの中身を取得
        td_list: bs4.element.ResultSet = soup.select('td')
        word_list: str = []

        # 単語を分割
        for item in td_list:
            word: str = splitter(item)
            word_list.append(word.split('<br/>'))

        # リストの空白要素を削除
        i: int = 0
        for element in word_list:
            if element[0] == '':
                word_list.pop(i)

            i += 1

        #Excelファイルに書き出す
        opr_excel.create_file(word_list)

        return

        # 単語を分割する関数
    def splitter(item: bs4.element.Tag):
        item: str = str(item)

        item = item.removeprefix(r'<td>')
        item = item.removesuffix(r'</td>')

        return item


class opr_excel():
    file_name: str = 'English_Word.xlsx'
    ws_name: str = 'word_data'

    # Excelファイルに書き出す
    def create_file(word_list: str):
        # ワークブックとワークシートを準備
        wb = openpyxl.Workbook()
        wb.active
        ws = wb.worksheets[0]

        # セルに値を代入する
        for i in range(len(word_list)):
            ws['A' + str(i+1)] = word_list[i][0]
            ws['B' + str(i+1)] = word_list[i][1]

        # シート名を変更
        ws.title = ws_name

        # 保存
        wb.save(file_name)

        print('ファイルの作成が終了しました。')

        return


if __name__ == '__main__':
    create_data_file.create_data()