import MySQLdb
import openpyxl

def main():
    # データベースに接続
    conn = MySQLdb.connect(
        user="***",
        passwd="***",
        host="localhost",
        db='english_word',
        charset='utf8'
    )

    cur = conn.cursor()

    # Excelファイルからデータを取得
    wb = openpyxl.load_workbook("./English_Word.xlsx")
    wb.active
    ws = wb["word_data"]
    max = ws.max_row


    for i in range(max):
        engWord = ws["A" + str(i+1)].value
        janWord = ws["B" + str(i+1)].value
    
        if engWord != "" and janWord != "":
            sql = f"insert word(eng, jan) values(\'{engWord}\', \'{janWord}\')"
            
            cur.execute(sql)
            conn.commit()

    
    cur.close()
    conn.close()

if __name__ == '__main__':
    main()